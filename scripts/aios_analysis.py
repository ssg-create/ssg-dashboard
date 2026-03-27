#!/usr/bin/env python3
"""
AIOS Analysis — Geração automática de aios-insights.json
Análise baseada em regras — sem API externa, zero custo.
"""

import os, json, glob, sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict

# ── Configuração ────────────────────────────────────────────
PASTA       = str(Path.home() / "Desktop" / "Arquivo OTRS")
OUTPUT_PATH = os.path.join(PASTA, "aios-insights.json")
LOG_PATH    = os.path.join(PASTA, "ssg_log.txt")

# ── Mapeamentos (espelham o dashboard) ──────────────────────
COL_MAP = {
    "num":      ["número do chamado", "número", "chamado", "ticket"],
    "criado":   ["criado", "data de abertura", "abertura", "created"],
    "fechado":  ["fechado", "data de fechamento", "fechamento", "closed"],
    "estado":   ["estado", "status", "situação"],
    "prior":    ["prioridade", "priority"],
    "atend":    ["atendente", "assignee"],
    "cli_nome": ["nome do cliente", "customer name", "c_cliente", "c cliente"],
    "assunto":  ["assunto", "subject"],
    "servico":  ["serviço", "service"],
    "resp_min": ["primeira resposta em minutos", "first response in minutes"],
    "sol_min":  ["tempo de solução em minutos", "resolution time in minutes"],
    "fila":     ["fila", "queue"],
}

ESTADO_MAP = {
    "fechado com êxito": "Fechado",
    "fechado sem êxito (auto)": "Fechado s/ Êxito",
    "fechado sem êxito": "Fechado s/ Êxito",
    "aguardando cliente": "Aguardando Cliente",
    "resolvido": "Resolvido",
    "aguardando atendente (interno)": "Aguardando Interno",
    "retorno cliente": "Retorno Cliente",
    "em atendimento": "Em Atendimento",
    "aberto": "Aberto",
    "aguardando agentes externos": "Aguardando Externo",
    "cancelado": "Fechado s/ Êxito",
    "novo": "Aberto",
}

PRIOR_MAP = {
    "1 very low": "Muito Baixa",
    "2 low": "Baixa",
    "3 normal": "Normal",
    "4 high": "Alta",
    "5 very high": "Muito Alta",
}

ABERTO    = {"Aberto", "Em Atendimento", "Aguardando Cliente", "Aguardando Interno", "Retorno Cliente", "Aguardando Externo"}
FECHADO   = {"Fechado", "Resolvido"}
BLOQUEADO = {"Aguardando Cliente", "Aguardando Interno", "Retorno Cliente", "Aguardando Externo"}

DIAS_PT = {
    "Monday": "segunda-feira", "Tuesday": "terça-feira", "Wednesday": "quarta-feira",
    "Thursday": "quinta-feira", "Friday": "sexta-feira", "Saturday": "sábado", "Sunday": "domingo"
}

# ── Helpers ─────────────────────────────────────────────────
def log(m):
    msg = f"[{datetime.now().strftime('%H:%M:%S')}] [AIOS] {m}"
    print(msg)
    try:
        with open(LOG_PATH, "a") as f:
            f.write(msg + "\n")
    except:
        pass

def find_col(headers, key):
    for h in headers:
        hl = h.lower().strip().replace("_", " ")
        for alias in COL_MAP.get(key, []):
            if hl == alias or alias in hl:
                return h
    return None

def ultimo_xlsx():
    arquivos = glob.glob(os.path.join(PASTA, "ticket_search_*.xlsx"))
    return sorted(arquivos)[-1] if arquivos else None

def mediana(lst):
    s = sorted(lst)
    n = len(s)
    return round(s[n // 2], 1) if n else 0

def media(lst):
    return round(sum(lst) / len(lst), 1) if lst else 0

# ── Leitura do xlsx ──────────────────────────────────────────
def processar_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        log("ERRO: openpyxl não instalado. Rode: pip3 install openpyxl --break-system-packages")
        sys.exit(1)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        log("ERRO: arquivo xlsx vazio")
        sys.exit(1)

    headers = [str(h or "").strip() for h in rows[0]]
    C   = {k: find_col(headers, k) for k in COL_MAP}
    idx = {h: i for i, h in enumerate(headers)}

    def get(row, key):
        col = C.get(key)
        if not col or col not in idx:
            return None
        return row[idx[col]]

    tickets = []
    for row in rows[1:]:
        if all(v is None or v == "" for v in row):
            continue

        estado = ESTADO_MAP.get(str(get(row, "estado") or "").lower().strip(), "Aberto")
        prior  = PRIOR_MAP.get(str(get(row, "prior") or "").lower().strip(), "Normal")

        try:
            resp_min = float(get(row, "resp_min") or 0) or None
        except:
            resp_min = None

        try:
            sol_h = round(float(get(row, "sol_min") or 0) / 60, 1) or None
        except:
            sol_h = None

        criado_raw = get(row, "criado")
        if isinstance(criado_raw, datetime):
            criado = criado_raw
        elif isinstance(criado_raw, str) and criado_raw.strip():
            try:
                criado = datetime.fromisoformat(criado_raw.strip().replace(" ", "T"))
            except:
                criado = None
        else:
            criado = None

        cli     = str(get(row, "cli_nome") or "").strip() or None
        servico = str(get(row, "servico") or "").split("::")[-1].strip() or "Outros"

        tickets.append({
            "estado": estado, "prior": prior,
            "cli": cli, "servico": servico,
            "resp_min": resp_min, "sol_h": sol_h,
            "criado": criado,
        })

    return tickets

# ── Métricas ─────────────────────────────────────────────────
def calcular_metricas(tickets):
    total    = len(tickets)
    now      = datetime.now()

    abertos    = [t for t in tickets if t["estado"] in ABERTO]
    fechados   = [t for t in tickets if t["estado"] in FECHADO]
    bloqueados = [t for t in tickets if t["estado"] in BLOQUEADO]

    prioridades = defaultdict(int)
    por_cliente = defaultdict(list)
    por_servico = defaultdict(int)
    por_dia     = defaultdict(int)
    por_hora    = defaultdict(int)

    resps, sols = [], []

    for t in tickets:
        prioridades[t["prior"]] += 1
        por_servico[t["servico"]] += 1
        if t["cli"]:
            por_cliente[t["cli"]].append(t)
        if t["resp_min"]:
            resps.append(t["resp_min"])
        if t["sol_h"]:
            sols.append(t["sol_h"])
        if t["criado"]:
            por_dia[t["criado"].strftime("%A")]  += 1
            por_hora[t["criado"].hour]            += 1

    # Aging tickets abertos
    aging = defaultdict(int)
    for t in abertos:
        if t["criado"]:
            d = (now - t["criado"]).days
            if   d <= 2:  aging["0_2"]  += 1
            elif d <= 7:  aging["3_7"]  += 1
            elif d <= 30: aging["8_30"] += 1
            else:         aging["30p"]  += 1

    resp_med = mediana(resps)
    sol_med  = media(sols)
    pct_ab   = round(len(abertos) / total * 100) if total else 0
    pct_ma   = round(prioridades["Muito Alta"] / total * 100) if total else 0

    # Health score (mesma lógica do dashboard)
    sc_resp = 100 if resp_med<=30 else 85 if resp_med<=45 else 70 if resp_med<=60 else 50 if resp_med<=90 else 30
    sc_sol  = 100 if sol_med<=24  else 85 if sol_med<=48  else 70 if sol_med<=60  else 50 if sol_med<=80  else 30
    nb      = len(bloqueados)
    sc_bloq = 100 if nb==0 else 90 if nb<=5 else 75 if nb<=10 else 55 if nb<=20 else 35 if nb<=40 else 15
    sc_ab   = 100 if pct_ab<=5 else 85 if pct_ab<=10 else 70 if pct_ab<=20 else 50 if pct_ab<=30 else 30
    health  = round(sc_resp*0.30 + sc_sol*0.25 + sc_bloq*0.25 + sc_ab*0.20)

    pico_dia  = DIAS_PT.get(max(por_dia,  key=por_dia.get),  "—") if por_dia  else "—"
    pico_hora = max(por_hora, key=por_hora.get) if por_hora else None

    top_clientes = sorted(por_cliente.items(), key=lambda x: -len(x[1]))[:5]
    top_servicos = sorted(por_servico.items(), key=lambda x: -x[1])[:5]

    return {
        "total": total, "abertos": len(abertos), "fechados": len(fechados),
        "bloqueados": nb, "pct_abertos": pct_ab, "pct_muito_alta": pct_ma,
        "taxa_conclusao": round(len(fechados) / total * 100) if total else 0,
        "prioridades": dict(prioridades),
        "resp_med": resp_med, "sol_med": sol_med,
        "pico_dia": pico_dia, "pico_hora": pico_hora,
        "aging": dict(aging), "health": health,
        "top_clientes": top_clientes, "top_servicos": top_servicos,
        "por_cliente": dict(por_cliente),
    }

# ── Geração de texto por regras ──────────────────────────────
def nivel_cor(n):
    return {"high": "var(--danger)", "medium": "var(--warn)", "low": "var(--ok)"}[n]

def nivel_saude(h):
    return "high" if h < 60 else "medium" if h < 80 else "low"

def gerar_narrativa_resumo(m):
    partes = []

    # Taxa de conclusão
    if m["taxa_conclusao"] >= 80:
        partes.append(f"A operação apresenta <strong>taxa de conclusão de {m['taxa_conclusao']}%</strong>, considerada excelente.")
    elif m["taxa_conclusao"] >= 65:
        partes.append(f"A operação apresenta <strong>taxa de conclusão de {m['taxa_conclusao']}%</strong>, com margem para melhoria.")
    else:
        partes.append(f"A taxa de conclusão está em <strong>{m['taxa_conclusao']}%</strong>, abaixo do esperado — requer atenção imediata.")

    # Abertos e bloqueados
    if m["pct_abertos"] > 25 and m["bloqueados"] > 15:
        partes.append(f"Existem <strong>{m['abertos']} chamados em aberto</strong> ({m['pct_abertos']}%) e <strong>{m['bloqueados']} bloqueados</strong> aguardando ação — combinação que indica gargalo operacional.")
    elif m["bloqueados"] > 15:
        partes.append(f"O volume de <strong>{m['bloqueados']} chamados bloqueados</strong> está crítico e exige triagem imediata.")
    elif m["pct_abertos"] > 20:
        partes.append(f"Com <strong>{m['pct_abertos']}% do volume em aberto</strong>, recomenda-se revisão do backlog.")

    # SLA
    if m["resp_med"] > 60:
        partes.append(f"O tempo médio de <strong>primeira resposta ({m['resp_med']} min)</strong> está bem acima da meta de 30 min — redistribuição de carga ou reforço de equipe é necessário.")
    elif m["resp_med"] > 30:
        partes.append(f"A <strong>primeira resposta ({m['resp_med']} min)</strong> está acima da meta de 30 min, mas controlável.")

    if m["sol_med"] > 60:
        partes.append(f"O tempo médio de <strong>solução de {m['sol_med']}h</strong> é elevado — revisar os chamados mais antigos deve ser prioridade.")

    # Pico
    if m["pico_dia"] != "—" and m["pico_hora"] is not None:
        partes.append(f"O pico de volume ocorre na <strong>{m['pico_dia']} às {m['pico_hora']}h</strong> — dimensionar equipe para esse intervalo pode reduzir o SLA médio.")

    # Top serviço
    if m["top_servicos"]:
        svc, cnt = m["top_servicos"][0]
        pct_svc = round(cnt / m["total"] * 100)
        if pct_svc > 20:
            partes.append(f"O módulo <strong>{svc}</strong> concentra <strong>{pct_svc}% dos chamados</strong> — avaliar capacitação técnica ou documentação de auto-atendimento.")

    return "<br><br>".join(partes)

def gerar_riscos_operacionais(m):
    riscos = []

    # SLA resposta
    if m["resp_med"] > 60:
        riscos.append({"titulo": "Tempo de primeira resposta crítico",
            "desc": f"Mediana de {m['resp_med']} min — o dobro da meta de 30 min. Clientes estão esperando mais de 1h pela primeira resposta, risco direto de insatisfação.",
            "nivel": "high", "confianca": 90})
    elif m["resp_med"] > 30:
        riscos.append({"titulo": "Tempo de primeira resposta acima da meta",
            "desc": f"Mediana de {m['resp_med']} min contra meta de 30 min. Tendência de deterioração se não houver intervenção.",
            "nivel": "medium", "confianca": 78})

    # Bloqueados
    if m["bloqueados"] > 30:
        riscos.append({"titulo": "Volume crítico de chamados bloqueados",
            "desc": f"{m['bloqueados']} chamados parados aguardando resposta ou ação interna. Representa gargalo operacional que impede o fechamento do backlog.",
            "nivel": "high", "confianca": 92})
    elif m["bloqueados"] > 15:
        riscos.append({"titulo": "Chamados bloqueados acima do normal",
            "desc": f"{m['bloqueados']} chamados em espera. Recomenda-se triagem para identificar os que dependem de ação interna e os que aguardam cliente.",
            "nivel": "medium", "confianca": 80})

    # Aging crítico
    aging_30 = m["aging"].get("30p", 0)
    aging_8_30 = m["aging"].get("8_30", 0)
    if aging_30 > 5:
        riscos.append({"titulo": f"{aging_30} chamados abertos há mais de 30 dias",
            "desc": f"Chamados com aging superior a 30 dias representam risco de escalonamento pelo cliente e impacto direto no health score da conta.",
            "nivel": "high", "confianca": 95})
    elif aging_8_30 > 10:
        riscos.append({"titulo": f"{aging_8_30} chamados entre 8 e 30 dias sem resolução",
            "desc": "Volume de chamados no limite antes de se tornarem críticos. Intervenção preventiva nos próximos 2 dias pode evitar escalonamento.",
            "nivel": "medium", "confianca": 82})

    # Prioridade muito alta
    if m["pct_muito_alta"] > 40:
        riscos.append({"titulo": "Concentração anormal de chamados críticos",
            "desc": f"{m['pct_muito_alta']}% dos chamados com prioridade Muito Alta — acima de 40% indica possível instabilidade sistêmica ou problema recorrente não resolvido.",
            "nivel": "high", "confianca": 75})
    elif m["pct_muito_alta"] > 28:
        riscos.append({"titulo": "Prioridade Muito Alta elevada",
            "desc": f"{m['pct_muito_alta']}% do volume em prioridade crítica. Monitorar se há padrão de módulos ou clientes específicos causando o aumento.",
            "nivel": "medium", "confianca": 70})

    # Concentração de horário
    if m["pico_dia"] != "—" and m["pico_hora"] is not None:
        por_dia_val = m["total"] // 5 if m["total"] > 0 else 1
        riscos.append({"titulo": f"Pico de volume na {m['pico_dia']} às {m['pico_hora']}h",
            "desc": f"Concentração identificada neste intervalo. Sem escalonamento preventivo, a equipe absorve o pico com capacidade regular, elevando o tempo de resposta.",
            "nivel": "low", "confianca": 65})

    # Módulo recorrente
    if m["top_servicos"]:
        svc, cnt = m["top_servicos"][0]
        pct = round(cnt / m["total"] * 100)
        if pct > 25:
            riscos.append({"titulo": f"Módulo {svc} concentra {pct}% dos chamados",
                "desc": f"Volume desproporcional no módulo {svc} sugere problema recorrente, gap de conhecimento dos usuários ou instabilidade técnica não endereçada.",
                "nivel": "medium", "confianca": 72})

    if not riscos:
        riscos.append({"titulo": "Operação dentro dos parâmetros",
            "desc": "Nenhum padrão de risco crítico identificado neste período. Manter monitoramento preventivo.",
            "nivel": "low", "confianca": 85})

    return riscos

def gerar_riscos_clientes(m):
    riscos_cli = []
    now = datetime.now()
    media_vol = m["total"] / max(len(m["por_cliente"]), 1)

    for cliente, tickets in m["top_clientes"]:
        vol = len(tickets)
        pct_vol = round(vol / m["total"] * 100)

        # Aging desta conta
        aging_30 = sum(1 for t in tickets if t["estado"] in ABERTO and t["criado"] and (now - t["criado"]).days > 30)
        aging_15 = sum(1 for t in tickets if t["estado"] in ABERTO and t["criado"] and 15 <= (now - t["criado"]).days <= 30)
        abertos_cli = sum(1 for t in tickets if t["estado"] in ABERTO)
        bloq_cli = sum(1 for t in tickets if t["estado"] in BLOQUEADO)

        # Prioridade crítica desta conta
        ma_cli = sum(1 for t in tickets if t["prior"] == "Muito Alta")
        pct_ma_cli = round(ma_cli / vol * 100) if vol else 0

        # SLA desta conta
        resps_cli = [t["resp_min"] for t in tickets if t["resp_min"]]
        resp_cli = mediana(resps_cli)

        # Serviço mais frequente desta conta
        svcs = defaultdict(int)
        for t in tickets:
            svcs[t["servico"]] += 1
        top_svc = max(svcs.items(), key=lambda x: x[1]) if svcs else ("—", 0)
        pct_svc = round(top_svc[1] / vol * 100) if vol else 0

        riscos = []
        narrativa_partes = []

        # Volume relativo
        if vol > media_vol * 1.5:
            narrativa_partes.append(f"Conta com <strong>volume acima da média</strong> — {vol} chamados ({pct_vol}% do total).")
            riscos.append({"titulo": "Volume acima da média do período",
                "desc": f"{vol} chamados ({pct_vol}% do total). Verificar se há problema sistêmico ou sazonalidade.",
                "nivel": "medium", "confianca": 78})

        if aging_30 > 0:
            narrativa_partes.append(f"<strong>{aging_30} chamado(s) aberto(s) há mais de 30 dias</strong> — risco imediato de escalonamento.")
            riscos.append({"titulo": f"{aging_30} chamado(s) com aging crítico (+30 dias)",
                "desc": "Chamados acumulados sem resolução representam risco direto de insatisfação e possível escalonamento para a gestão do cliente.",
                "nivel": "high", "confianca": 93})

        if aging_15 > 2:
            narrativa_partes.append(f"{aging_15} chamados entre 15 e 30 dias sem resolução — intervenção preventiva recomendada.")
            riscos.append({"titulo": f"{aging_15} chamados próximos do limite de aging",
                "desc": "Chamados entre 15–30 dias que podem virar críticos. Priorização nos próximos 2 dias evita escalonamento.",
                "nivel": "medium", "confianca": 80})

        if resp_cli > 60:
            riscos.append({"titulo": f"SLA de resposta fora do contrato ({resp_cli} min)",
                "desc": f"Mediana de primeira resposta nesta conta: {resp_cli} min. Acima da meta contratual de 30 min.",
                "nivel": "high" if resp_cli > 90 else "medium", "confianca": 85})

        if pct_ma_cli > 40:
            narrativa_partes.append(f"<strong>{pct_ma_cli}% dos chamados com prioridade crítica</strong> — padrão acima do normal.")
            riscos.append({"titulo": f"Alto percentual de prioridade Muito Alta ({pct_ma_cli}%)",
                "desc": "Pode indicar instabilidade no ambiente do cliente ou processo de abertura de chamados com triagem inadequada.",
                "nivel": "medium", "confianca": 70})

        if pct_svc > 40:
            narrativa_partes.append(f"O módulo <strong>{top_svc[0]}</strong> concentra {pct_svc}% dos chamados desta conta — possível reincidência.")
            riscos.append({"titulo": f"Reincidência no módulo {top_svc[0]}",
                "desc": f"{pct_svc}% dos chamados desta conta são do mesmo módulo. Sugere problema não resolvido na raiz ou gap de treinamento.",
                "nivel": "medium", "confianca": 75})

        if not riscos:
            riscos.append({"titulo": "Conta dentro dos parâmetros normais",
                "desc": "Sem padrões de risco identificados para este cliente neste período.",
                "nivel": "low", "confianca": 80})
            narrativa_partes.append("Conta operando dentro dos parâmetros esperados para o período.")

        narrativa = " ".join(narrativa_partes) if narrativa_partes else f"Conta com {vol} chamados no período, sem anomalias críticas identificadas."

        riscos_cli.append({
            "cliente": cliente,
            "narrativa": narrativa,
            "riscos": riscos[:4]  # máx 4 riscos por cliente
        })

    return riscos_cli

def gerar_cs_insights(m):
    insights = []
    now = datetime.now()

    for cliente, tickets in m["top_clientes"][:3]:  # top 3 contas
        vol = len(tickets)
        abertos_cli = [t for t in tickets if t["estado"] in ABERTO]
        aging_30 = [t for t in abertos_cli if t["criado"] and (now - t["criado"]).days > 30]
        aging_15 = [t for t in abertos_cli if t["criado"] and 15 <= (now - t["criado"]).days <= 30]

        svcs = defaultdict(int)
        for t in tickets:
            svcs[t["servico"]] += 1
        top_svc = max(svcs.items(), key=lambda x: x[1]) if svcs else ("—", 0)
        pct_svc = round(top_svc[1] / vol * 100) if vol else 0

        resps_cli = [t["resp_min"] for t in tickets if t["resp_min"]]
        resp_cli = mediana(resps_cli)

        partes = []
        proximo_contato = "nos próximos 7 dias"
        tom = "consultivo"
        prioridade = "baixa"

        # Aging crítico → contato urgente
        if aging_30:
            partes.append(f"Conta com <strong>{len(aging_30)} chamado(s) aberto(s) há mais de 30 dias</strong>. Contato proativo urgente necessário antes que o cliente acione a gestão.")
            proximo_contato = "em até 2 dias úteis"
            tom = "proativo · urgente"
            prioridade = "alta"
        elif aging_15:
            partes.append(f"<strong>{len(aging_15)} chamado(s) próximos do limite de 30 dias</strong> sem resolução. Contato preventivo recomendado para demonstrar atenção antes do vencimento.")
            proximo_contato = "até o fim desta semana"
            tom = "proativo"
            prioridade = "média"

        # Reincidência de módulo
        if pct_svc > 40:
            partes.append(f"O módulo <strong>{top_svc[0]}</strong> representa {pct_svc}% dos chamados desta conta. Sugerir uma <strong>sessão de alinhamento técnico</strong> pode quebrar o ciclo de reincidência e reduzir volume.")
            if prioridade == "baixa":
                prioridade = "média"
                tom = "consultivo"

        # SLA fora
        if resp_cli > 60:
            partes.append(f"O tempo médio de resposta para esta conta foi de <strong>{resp_cli} min</strong>, acima do SLA contratado. Reconhecer proativamente e apresentar plano de melhoria.")
            if prioridade != "alta":
                prioridade = "média"
            tom = "proativo · transparente"

        # Saúde geral
        pct_ab_cli = round(len(abertos_cli) / vol * 100) if vol else 0
        if pct_ab_cli > 30:
            partes.append(f"Com <strong>{pct_ab_cli}% dos chamados ainda em aberto</strong>, o próximo contato deve incluir um plano de fechamento claro com prazo.")
        elif not partes:
            partes.append(f"Conta com {vol} chamados no período e sem anomalias críticas. Manter relacionamento regular para garantir satisfação.")
            tom = "consultivo"

        narrativa = "<br><br>".join(partes)

        insights.append({
            "cliente": cliente,
            "narrativa": narrativa,
            "proximo_contato": proximo_contato,
            "tom_sugerido": tom,
            "prioridade": prioridade,
        })

    return insights

# ── Montagem final do JSON ───────────────────────────────────
def gerar_insights(m):
    health = m["health"]
    health_cor = "var(--ok)" if health >= 80 else "var(--warn)" if health >= 60 else "var(--danger)"
    risco_sla = "Alto" if m["resp_med"] > 60 else "Médio" if m["resp_med"] > 30 else "Baixo"
    risco_sla_cor = "var(--danger)" if risco_sla == "Alto" else "var(--warn)" if risco_sla == "Médio" else "var(--ok)"

    # Previsão simples baseada na média dos últimos dados
    prev_dia = round(m["total"] / 53) if m["total"] > 0 else 0  # ~53 dias úteis em jan-mar

    top_svc_nome = m["top_servicos"][0][0] if m["top_servicos"] else "—"
    oport_txt = f"-{min(15, round(m['pct_muito_alta'] * 0.3))}% com triagem" if m["pct_muito_alta"] > 20 else "Capacitação recomendada"
    oport_desc = f"Foco no módulo {top_svc_nome} pode reduzir reincidências" if top_svc_nome != "—" else "Revisar processos de triagem"

    risco_principal_txt = f"{m['aging'].get('30p', 0)} chamados +30 dias" if m["aging"].get("30p", 0) > 0 else f"SLA {m['resp_med']} min"
    risco_desc = "Chamados críticos sem resolução" if m["aging"].get("30p", 0) > 0 else "Acima da meta de 30 min"

    return {
        "generated_at": datetime.now().isoformat(),
        "version": "1.0",
        "resumo_gestao": {
            "narrativa": gerar_narrativa_resumo(m),
            "kpis": [
                {"label": "Score Operacional", "value": str(health),
                 "sub": f"{'excelente' if health>=80 else 'atenção' if health>=60 else 'crítico'} · meta: >80", "color": health_cor},
                {"label": "Risco de SLA", "value": risco_sla,
                 "sub": f"1ª resposta: {m['resp_med']} min", "color": risco_sla_cor},
                {"label": "Oportunidade", "value": oport_txt,
                 "sub": oport_desc, "color": "var(--teal)"},
                {"label": "Previsão 7 dias", "value": f"~{prev_dia * 5}/semana",
                 "sub": f"pico: {m['pico_dia']} {m['pico_hora']}h" if m["pico_hora"] is not None else "baseado no ritmo atual",
                 "color": "var(--purple)"},
            ],
            "risco_principal": {"label": risco_principal_txt, "desc": risco_desc},
            "oportunidade":    {"label": oport_txt,           "desc": oport_desc},
            "previsao":        {"label": f"~{prev_dia}/dia",  "desc": f"pico: {m['pico_dia']} {m['pico_hora']}h" if m["pico_hora"] is not None else "baseado no ritmo atual"},
        },
        "riscos_operacionais": gerar_riscos_operacionais(m),
        "riscos_clientes":     gerar_riscos_clientes(m),
        "cs_insights":         gerar_cs_insights(m),
    }

# ── Main ─────────────────────────────────────────────────────
def main():
    log("Iniciando análise AIOS (modo regras)...")

    xlsx_path = ultimo_xlsx()
    if not xlsx_path:
        log("ERRO: nenhum arquivo ticket_search_*.xlsx encontrado em " + PASTA)
        sys.exit(1)

    log(f"Processando: {os.path.basename(xlsx_path)}")
    tickets = processar_xlsx(xlsx_path)
    log(f"{len(tickets)} chamados carregados")

    m = calcular_metricas(tickets)
    log(f"health={m['health']} | abertos={m['abertos']} | bloqueados={m['bloqueados']} | resp={m['resp_med']}min | sol={m['sol_med']}h")

    insights = gerar_insights(m)

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(insights, f, ensure_ascii=False, indent=2)

    log(f"aios-insights.json salvo — {len(insights['riscos_operacionais'])} riscos op · {len(insights['riscos_clientes'])} clientes · {len(insights['cs_insights'])} insights CS")
    log("Análise AIOS concluída.")

if __name__ == "__main__":
    main()
